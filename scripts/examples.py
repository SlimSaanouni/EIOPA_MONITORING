"""
Exemples d'utilisation avancée du système de monitoring EIOPA
"""
import pandas as pd
from datetime import datetime, timedelta
from pathlib import Path

from src.analyzer import EIOPAAnalyzer
from src.downloader import EIOPADownloader
from src.ingestion import ingest_zip
from src.utils import format_rate_pct, format_bps


def example_1_basic_usage():
    """
    Exemple 1 : Utilisation basique - Télécharger et analyser
    """
    print("\n" + "=" * 80)
    print("EXEMPLE 1 : Utilisation basique")
    print("=" * 80 + "\n")
    
    # Télécharger le dernier fichier
    downloader = EIOPADownloader()
    zip_path = downloader.download_latest()
    
    if not zip_path:
        print("❌ Échec du téléchargement")
        return
    
    # Ingérer (extraction + écriture dans historical.db)
    try:
        data = ingest_zip(zip_path)
    except ValueError as e:
        print(f"❌ Échec de l'ingestion : {e}")
        return
    
    # Afficher les résultats
    print(f"✅ Données extraites pour {data['reference_date'].strftime('%Y-%m-%d')}")
    print("\nTaux EUR (France) :")
    for maturity, rate in sorted(data['rates'].items()):
        print(f"  {maturity:2d}Y : {format_rate_pct(rate)}")
    
    if data.get('va'):
        print(f"\nVolatility Adjustment : {format_rate_pct(data['va'])}")


def example_2_historical_analysis():
    """
    Exemple 2 : Analyse historique et séries temporelles
    """
    print("\n" + "=" * 80)
    print("EXEMPLE 2 : Analyse historique")
    print("=" * 80 + "\n")
    
    analyzer = EIOPAAnalyzer()
    
    if analyzer.historical_data.empty:
        print("⚠️ Pas d'historique disponible. Exécutez d'abord example_1 plusieurs fois.")
        return
    
    # Statistiques globales
    print(f"Nombre d'enregistrements : {len(analyzer.historical_data)}")
    print(f"Période : {analyzer.historical_data['reference_date'].min().strftime('%Y-%m-%d')} "
          f"à {analyzer.historical_data['reference_date'].max().strftime('%Y-%m-%d')}")
    
    # Série temporelle taux 10Y
    print("\n📈 Évolution du taux 10Y (6 derniers mois) :")
    
    six_months_ago = datetime.now() - timedelta(days=180)
    ts = analyzer.get_time_series('FR', 10, start_date=six_months_ago)
    
    if not ts.empty:
        print("\nDate          | Taux 10Y")
        print("-" * 30)
        for _, row in ts.iterrows():
            print(f"{row['reference_date'].strftime('%Y-%m-%d')} | {format_rate_pct(row['rate'])}")
        
        # Statistiques
        print(f"\nMin : {format_rate_pct(ts['rate'].min())}")
        print(f"Max : {format_rate_pct(ts['rate'].max())}")
        print(f"Moy : {format_rate_pct(ts['rate'].mean())}")


def example_3_multi_maturity_comparison():
    """
    Exemple 3 : Comparaison de plusieurs maturités
    """
    print("\n" + "=" * 80)
    print("EXEMPLE 3 : Comparaison de maturités")
    print("=" * 80 + "\n")
    
    analyzer = EIOPAAnalyzer()
    
    if analyzer.historical_data.empty:
        print("⚠️ Pas d'historique disponible")
        return
    
    # Dernières données
    latest = analyzer.historical_data.iloc[-1]
    date = latest['reference_date']
    
    print(f"Date : {date.strftime('%Y-%m-%d')}\n")
    
    # Calculer les spreads
    print("Spreads de taux :")
    print("-" * 40)
    
    rate_1y = latest['rate_1y']
    rate_10y = latest['rate_10y']
    rate_30y = latest['rate_30y']
    
    spread_10_1 = (rate_10y - rate_1y) * 10000
    spread_30_10 = (rate_30y - rate_10y) * 10000
    spread_30_1 = (rate_30y - rate_1y) * 10000
    
    print(f"10Y - 1Y  : {format_bps(spread_10_1)}")
    print(f"30Y - 10Y : {format_bps(spread_30_10)}")
    print(f"30Y - 1Y  : {format_bps(spread_30_1)}")
    
    # Pente de la courbe
    print(f"\n📐 Pente moyenne : {format_bps((rate_30y - rate_1y) / 29 * 10000)} par an")


def example_4_custom_alerts():
    """
    Exemple 4 : Système d'alertes personnalisé
    """
    print("\n" + "=" * 80)
    print("EXEMPLE 4 : Alertes personnalisées")
    print("=" * 80 + "\n")
    
    analyzer = EIOPAAnalyzer()
    
    if len(analyzer.historical_data) < 2:
        print("⚠️ Pas assez de données pour la comparaison")
        return
    
    # Comparer les 2 dernières entrées
    latest = analyzer.historical_data.iloc[-1]
    previous = analyzer.historical_data.iloc[-2]
    
    print("Comparaison :")
    print(f"  Actuel   : {latest['reference_date'].strftime('%Y-%m-%d')}")
    print(f"  Précédent : {previous['reference_date'].strftime('%Y-%m-%d')}")
    print()
    
    # Seuils personnalisés
    ALERT_THRESHOLD_CRITICAL = 100  # bps
    ALERT_THRESHOLD_WARNING = 50    # bps
    
    alerts = []
    
    for maturity in [1, 5, 10, 20, 30]:
        col = f'rate_{maturity}y'
        
        if pd.notna(latest[col]) and pd.notna(previous[col]):
            change_bps = (latest[col] - previous[col]) * 10000
            
            if abs(change_bps) >= ALERT_THRESHOLD_CRITICAL:
                level = "🔴 CRITIQUE"
                alerts.append((level, maturity, change_bps))
            elif abs(change_bps) >= ALERT_THRESHOLD_WARNING:
                level = "🟠 ATTENTION"
                alerts.append((level, maturity, change_bps))
    
    if alerts:
        print("Alertes détectées :")
        for level, maturity, change in alerts:
            direction = "hausse" if change > 0 else "baisse"
            print(f"  {level} - Taux {maturity}Y : {direction} de {format_bps(change)}")
    else:
        print("✅ Aucune alerte, variations normales")


def example_5_export_to_excel():
    """
    Exemple 5 : Export Excel avancé avec formatting
    """
    print("\n" + "=" * 80)
    print("EXEMPLE 5 : Export Excel avancé")
    print("=" * 80 + "\n")
    
    analyzer = EIOPAAnalyzer()
    
    if analyzer.historical_data.empty:
        print("⚠️ Pas de données à exporter")
        return
    
    output_file = Path("data/eiopa_export_avance.xlsx")
    
    try:
        from openpyxl.styles import Font, PatternFill, Alignment
        
        with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
            # Feuille 1 : Données brutes
            df_raw = analyzer.historical_data.copy()
            df_raw.to_excel(writer, sheet_name='Données brutes', index=False)
            
            # Feuille 2 : Taux 10Y uniquement
            df_10y = df_raw[['reference_date', 'rate_10y']].copy()
            df_10y.columns = ['Date', 'Taux 10Y (%)']
            df_10y['Taux 10Y (%)'] = df_10y['Taux 10Y (%)'] * 100
            df_10y.to_excel(writer, sheet_name='Taux 10Y', index=False)
            
            # Feuille 3 : Variations mensuelles
            if len(df_raw) > 1:
                variations = []
                for i in range(1, len(df_raw)):
                    current = df_raw.iloc[i]
                    previous = df_raw.iloc[i-1]
                    
                    row = {'Date': current['reference_date']}
                    
                    for maturity in [1, 5, 10, 20, 30]:
                        col = f'rate_{maturity}y'
                        if pd.notna(current[col]) and pd.notna(previous[col]):
                            change = (current[col] - previous[col]) * 10000
                            row[f'Δ {maturity}Y (bps)'] = change
                    
                    variations.append(row)
                
                df_variations = pd.DataFrame(variations)
                df_variations.to_excel(writer, sheet_name='Variations', index=False)
            
            # Formatter les en-têtes
            for sheet_name in writer.sheets:
                ws = writer.sheets[sheet_name]
                for cell in ws[1]:
                    cell.font = Font(bold=True)
                    cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
                    cell.font = Font(color="FFFFFF", bold=True)
                    cell.alignment = Alignment(horizontal='center')
        
        print(f"✅ Fichier Excel créé : {output_file}")
        
    except ImportError:
        print("⚠️ openpyxl non installé. Installation : pip install openpyxl")


def example_6_calculate_duration():
    """
    Exemple 6 : Calcul de la duration d'une courbe
    """
    print("\n" + "=" * 80)
    print("EXEMPLE 6 : Calcul de duration")
    print("=" * 80 + "\n")
    
    analyzer = EIOPAAnalyzer()
    
    if analyzer.historical_data.empty:
        print("⚠️ Pas de données disponibles")
        return
    
    latest = analyzer.historical_data.iloc[-1]
    
    # Extraire les taux
    rates = {}
    for maturity in [1, 5, 10, 20, 30]:
        col = f'rate_{maturity}y'
        if pd.notna(latest[col]):
            rates[maturity] = float(latest[col])
    
    if not rates:
        print("⚠️ Pas de taux disponibles")
        return
    
    print(f"Date : {latest['reference_date'].strftime('%Y-%m-%d')}\n")
    print("Calcul de duration d'un portefeuille obligataire fictif :")
    print("-" * 60)
    
    # Portefeuille exemple : 100M€ répartis sur différentes maturités
    portfolio = {
        1: 10,   # 10M€ à 1 an
        5: 25,   # 25M€ à 5 ans
        10: 40,  # 40M€ à 10 ans
        20: 20,  # 20M€ à 20 ans
        30: 5    # 5M€ à 30 ans
    }
    
    total_value = sum(portfolio.values())
    weighted_duration = 0
    
    print(f"\nRépartition du portefeuille (Total : {total_value}M€) :")
    print(f"{'Maturité':<10} | {'Montant':>10} | {'Taux':>8} | {'Duration*':>10}")
    print("-" * 50)
    
    for maturity, amount in sorted(portfolio.items()):
        if maturity in rates:
            rate = rates[maturity]
            # Duration simple (approximation)
            duration = maturity / (1 + rate)
            weight = amount / total_value
            weighted_duration += duration * weight
            
            print(f"{maturity:2d} ans     | {amount:>7.0f}M€ | {rate*100:>6.2f}% | {duration:>8.2f}")
    
    print("-" * 50)
    print(f"\n📊 Duration moyenne du portefeuille : {weighted_duration:.2f} ans")
    print(f"   Sensibilité : -1% de taux → +{weighted_duration:.2f}% de valeur")


def example_7_stress_testing():
    """
    Exemple 7 : Stress testing (chocs de taux)
    """
    print("\n" + "=" * 80)
    print("EXEMPLE 7 : Stress testing")
    print("=" * 80 + "\n")
    
    analyzer = EIOPAAnalyzer()
    
    if analyzer.historical_data.empty:
        print("⚠️ Pas de données disponibles")
        return
    
    latest = analyzer.historical_data.iloc[-1]
    
    # Scénarios de choc
    scenarios = {
        "Hausse parallèle +100bps": 0.01,
        "Hausse parallèle +200bps": 0.02,
        "Baisse parallèle -50bps": -0.005,
        "Aplatissement (courte +50, longue -50)": None,  # Spécial
        "Pentification (courte -50, longue +50)": None   # Spécial
    }
    
    print(f"Date de référence : {latest['reference_date'].strftime('%Y-%m-%d')}\n")
    print("Taux de base :")
    
    base_rates = {}
    for maturity in [1, 10, 30]:
        col = f'rate_{maturity}y'
        if pd.notna(latest[col]):
            base_rates[maturity] = float(latest[col])
            print(f"  {maturity:2d}Y : {format_rate_pct(base_rates[maturity])}")
    
    print("\n" + "-" * 60)
    print("Scénarios de stress :")
    print("-" * 60)
    
    for scenario_name, shock in scenarios.items():
        print(f"\n{scenario_name} :")
        
        for maturity in [1, 10, 30]:
            if maturity in base_rates:
                base = base_rates[maturity]
                
                if shock is not None:
                    # Choc parallèle
                    stressed = base + shock
                else:
                    # Chocs spéciaux
                    if "Aplatissement" in scenario_name:
                        stressed = base + 0.005 if maturity == 1 else base - 0.005
                    else:  # Pentification
                        stressed = base - 0.005 if maturity == 1 else base + 0.005
                
                change = (stressed - base) * 10000
                print(f"  {maturity:2d}Y : {format_rate_pct(base)} → {format_rate_pct(stressed)} ({format_bps(change)})")


def run_all_examples():
    """Exécute tous les exemples"""
    examples = [
        example_1_basic_usage,
        example_2_historical_analysis,
        example_3_multi_maturity_comparison,
        example_4_custom_alerts,
        example_5_export_to_excel,
        example_6_calculate_duration,
        example_7_stress_testing
    ]
    
    print("\n" + "🚀" * 40)
    print("EXÉCUTION DE TOUS LES EXEMPLES")
    print("🚀" * 40)
    
    for i, example in enumerate(examples, 1):
        try:
            example()
            input(f"\n▶️  Appuyez sur Entrée pour continuer vers l'exemple {i+1}..." 
                  if i < len(examples) else "\n✅ Tous les exemples terminés. Appuyez sur Entrée...")
        except Exception as e:
            print(f"\n❌ Erreur dans l'exemple {i} : {e}")


if __name__ == "__main__":
    import sys
    
    if len(sys.argv) > 1:
        example_num = sys.argv[1]
        
        examples_map = {
            '1': example_1_basic_usage,
            '2': example_2_historical_analysis,
            '3': example_3_multi_maturity_comparison,
            '4': example_4_custom_alerts,
            '5': example_5_export_to_excel,
            '6': example_6_calculate_duration,
            '7': example_7_stress_testing,
            'all': run_all_examples
        }
        
        if example_num in examples_map:
            examples_map[example_num]()
        else:
            print("Usage : python examples.py [1-7|all]")
    else:
        run_all_examples()