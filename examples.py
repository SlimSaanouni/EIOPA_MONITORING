"""
Exemples d'utilisation avanc√©e du syst√®me de monitoring EIOPA
"""
import pandas as pd
import numpy as np
from datetime import datetime, timedelta
from pathlib import Path

from analyzer import EIOPAAnalyzer
from downloader import EIOPADownloader
from processor import EIOPAProcessor
from utils import format_rate_pct, format_bps


def example_1_basic_usage():
    """
    Exemple 1 : Utilisation basique - T√©l√©charger et analyser
    """
    print("\n" + "=" * 80)
    print("EXEMPLE 1 : Utilisation basique")
    print("=" * 80 + "\n")
    
    # T√©l√©charger le dernier fichier
    downloader = EIOPADownloader()
    zip_path = downloader.download_latest()
    
    if not zip_path:
        print("‚ùå √âchec du t√©l√©chargement")
        return
    
    # Traiter
    processor = EIOPAProcessor(zip_path)
    data = processor.process()
    
    if not data:
        print("‚ùå √âchec du traitement")
        return
    
    # Afficher les r√©sultats
    print(f"‚úÖ Donn√©es extraites pour {data['reference_date'].strftime('%Y-%m-%d')}")
    print(f"\nTaux EUR (France) :")
    for maturity, rate in sorted(data['rates'].items()):
        print(f"  {maturity:2d}Y : {format_rate_pct(rate)}")
    
    if data.get('va'):
        print(f"\nVolatility Adjustment : {format_rate_pct(data['va'])}")


def example_2_historical_analysis():
    """
    Exemple 2 : Analyse historique et s√©ries temporelles
    """
    print("\n" + "=" * 80)
    print("EXEMPLE 2 : Analyse historique")
    print("=" * 80 + "\n")
    
    analyzer = EIOPAAnalyzer()
    
    if analyzer.historical_data.empty:
        print("‚ö†Ô∏è Pas d'historique disponible. Ex√©cutez d'abord example_1 plusieurs fois.")
        return
    
    # Statistiques globales
    print(f"Nombre d'enregistrements : {len(analyzer.historical_data)}")
    print(f"P√©riode : {analyzer.historical_data['reference_date'].min().strftime('%Y-%m-%d')} "
          f"√† {analyzer.historical_data['reference_date'].max().strftime('%Y-%m-%d')}")
    
    # S√©rie temporelle taux 10Y
    print("\nüìà √âvolution du taux 10Y (6 derniers mois) :")
    
    six_months_ago = datetime.now() - timedelta(days=180)
    ts = analyzer.get_time_series('FR', 10, start_date=six_months_ago)
    
    if not ts.empty:
        print("\nDate          | Taux 10Y")
        print("-" * 30)
        for _, row in ts.iterrows():
            print(f"{row['reference_date'].strftime('%Y-%m-%d')} | {format_rate_pct(row['rate'])}")
        
        # Statistiques
        print(f"\nMin : {format_rate_pct(ts['rate'].min())}")
        print(f"Max : {format_rate_pct(ts['rate'].max())}")
        print(f"Moy : {format_rate_pct(ts['rate'].mean())}")


def example_3_multi_maturity_comparison():
    """
    Exemple 3 : Comparaison de plusieurs maturit√©s
    """
    print("\n" + "=" * 80)
    print("EXEMPLE 3 : Comparaison de maturit√©s")
    print("=" * 80 + "\n")
    
    analyzer = EIOPAAnalyzer()
    
    if analyzer.historical_data.empty:
        print("‚ö†Ô∏è Pas d'historique disponible")
        return
    
    # Derni√®res donn√©es
    latest = analyzer.historical_data.iloc[-1]
    date = latest['reference_date']
    
    print(f"Date : {date.strftime('%Y-%m-%d')}\n")
    
    # Calculer les spreads
    print("Spreads de taux :")
    print("-" * 40)
    
    rate_1y = latest['rate_1y']
    rate_10y = latest['rate_10y']
    rate_30y = latest['rate_30y']
    
    spread_10_1 = (rate_10y - rate_1y) * 10000
    spread_30_10 = (rate_30y - rate_10y) * 10000
    spread_30_1 = (rate_30y - rate_1y) * 10000
    
    print(f"10Y - 1Y  : {format_bps(spread_10_1)}")
    print(f"30Y - 10Y : {format_bps(spread_30_10)}")
    print(f"30Y - 1Y  : {format_bps(spread_30_1)}")
    
    # Pente de la courbe
    print(f"\nüìê Pente moyenne : {format_bps((rate_30y - rate_1y) / 29 * 10000)} par an")


def example_4_custom_alerts():
    """
    Exemple 4 : Syst√®me d'alertes personnalis√©
    """
    print("\n" + "=" * 80)
    print("EXEMPLE 4 : Alertes personnalis√©es")
    print("=" * 80 + "\n")
    
    analyzer = EIOPAAnalyzer()
    
    if len(analyzer.historical_data) < 2:
        print("‚ö†Ô∏è Pas assez de donn√©es pour la comparaison")
        return
    
    # Comparer les 2 derni√®res entr√©es
    latest = analyzer.historical_data.iloc[-1]
    previous = analyzer.historical_data.iloc[-2]
    
    print(f"Comparaison :")
    print(f"  Actuel   : {latest['reference_date'].strftime('%Y-%m-%d')}")
    print(f"  Pr√©c√©dent : {previous['reference_date'].strftime('%Y-%m-%d')}")
    print()
    
    # Seuils personnalis√©s
    ALERT_THRESHOLD_CRITICAL = 100  # bps
    ALERT_THRESHOLD_WARNING = 50    # bps
    
    alerts = []
    
    for maturity in [1, 5, 10, 20, 30]:
        col = f'rate_{maturity}y'
        
        if pd.notna(latest[col]) and pd.notna(previous[col]):
            change_bps = (latest[col] - previous[col]) * 10000
            
            if abs(change_bps) >= ALERT_THRESHOLD_CRITICAL:
                level = "üî¥ CRITIQUE"
                alerts.append((level, maturity, change_bps))
            elif abs(change_bps) >= ALERT_THRESHOLD_WARNING:
                level = "üü† ATTENTION"
                alerts.append((level, maturity, change_bps))
    
    if alerts:
        print("Alertes d√©tect√©es :")
        for level, maturity, change in alerts:
            direction = "hausse" if change > 0 else "baisse"
            print(f"  {level} - Taux {maturity}Y : {direction} de {format_bps(change)}")
    else:
        print("‚úÖ Aucune alerte, variations normales")


def example_5_export_to_excel():
    """
    Exemple 5 : Export Excel avanc√© avec formatting
    """
    print("\n" + "=" * 80)
    print("EXEMPLE 5 : Export Excel avanc√©")
    print("=" * 80 + "\n")
    
    analyzer = EIOPAAnalyzer()
    
    if analyzer.historical_data.empty:
        print("‚ö†Ô∏è Pas de donn√©es √† exporter")
        return
    
    output_file = Path("data/eiopa_export_avance.xlsx")
    
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
        
        with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
            # Feuille 1 : Donn√©es brutes
            df_raw = analyzer.historical_data.copy()
            df_raw.to_excel(writer, sheet_name='Donn√©es brutes', index=False)
            
            # Feuille 2 : Taux 10Y uniquement
            df_10y = df_raw[['reference_date', 'rate_10y']].copy()
            df_10y.columns = ['Date', 'Taux 10Y (%)']
            df_10y['Taux 10Y (%)'] = df_10y['Taux 10Y (%)'] * 100
            df_10y.to_excel(writer, sheet_name='Taux 10Y', index=False)
            
            # Feuille 3 : Variations mensuelles
            if len(df_raw) > 1:
                variations = []
                for i in range(1, len(df_raw)):
                    current = df_raw.iloc[i]
                    previous = df_raw.iloc[i-1]
                    
                    row = {'Date': current['reference_date']}
                    
                    for maturity in [1, 5, 10, 20, 30]:
                        col = f'rate_{maturity}y'
                        if pd.notna(current[col]) and pd.notna(previous[col]):
                            change = (current[col] - previous[col]) * 10000
                            row[f'Œî {maturity}Y (bps)'] = change
                    
                    variations.append(row)
                
                df_variations = pd.DataFrame(variations)
                df_variations.to_excel(writer, sheet_name='Variations', index=False)
            
            # Formatter les en-t√™tes
            for sheet_name in writer.sheets:
                ws = writer.sheets[sheet_name]
                for cell in ws[1]:
                    cell.font = Font(bold=True)
                    cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
                    cell.font = Font(color="FFFFFF", bold=True)
                    cell.alignment = Alignment(horizontal='center')
        
        print(f"‚úÖ Fichier Excel cr√©√© : {output_file}")
        
    except ImportError:
        print("‚ö†Ô∏è openpyxl non install√©. Installation : pip install openpyxl")


def example_6_calculate_duration():
    """
    Exemple 6 : Calcul de la duration d'une courbe
    """
    print("\n" + "=" * 80)
    print("EXEMPLE 6 : Calcul de duration")
    print("=" * 80 + "\n")
    
    analyzer = EIOPAAnalyzer()
    
    if analyzer.historical_data.empty:
        print("‚ö†Ô∏è Pas de donn√©es disponibles")
        return
    
    latest = analyzer.historical_data.iloc[-1]
    
    # Extraire les taux
    rates = {}
    for maturity in [1, 5, 10, 20, 30]:
        col = f'rate_{maturity}y'
        if pd.notna(latest[col]):
            rates[maturity] = float(latest[col])
    
    if not rates:
        print("‚ö†Ô∏è Pas de taux disponibles")
        return
    
    print(f"Date : {latest['reference_date'].strftime('%Y-%m-%d')}\n")
    print("Calcul de duration d'un portefeuille obligataire fictif :")
    print("-" * 60)
    
    # Portefeuille exemple : 100M‚Ç¨ r√©partis sur diff√©rentes maturit√©s
    portfolio = {
        1: 10,   # 10M‚Ç¨ √† 1 an
        5: 25,   # 25M‚Ç¨ √† 5 ans
        10: 40,  # 40M‚Ç¨ √† 10 ans
        20: 20,  # 20M‚Ç¨ √† 20 ans
        30: 5    # 5M‚Ç¨ √† 30 ans
    }
    
    total_value = sum(portfolio.values())
    weighted_duration = 0
    
    print(f"\nR√©partition du portefeuille (Total : {total_value}M‚Ç¨) :")
    print(f"{'Maturit√©':<10} | {'Montant':>10} | {'Taux':>8} | {'Duration*':>10}")
    print("-" * 50)
    
    for maturity, amount in sorted(portfolio.items()):
        if maturity in rates:
            rate = rates[maturity]
            # Duration simple (approximation)
            duration = maturity / (1 + rate)
            weight = amount / total_value
            weighted_duration += duration * weight
            
            print(f"{maturity:2d} ans     | {amount:>7.0f}M‚Ç¨ | {rate*100:>6.2f}% | {duration:>8.2f}")
    
    print("-" * 50)
    print(f"\nüìä Duration moyenne du portefeuille : {weighted_duration:.2f} ans")
    print(f"   Sensibilit√© : -1% de taux ‚Üí +{weighted_duration:.2f}% de valeur")


def example_7_stress_testing():
    """
    Exemple 7 : Stress testing (chocs de taux)
    """
    print("\n" + "=" * 80)
    print("EXEMPLE 7 : Stress testing")
    print("=" * 80 + "\n")
    
    analyzer = EIOPAAnalyzer()
    
    if analyzer.historical_data.empty:
        print("‚ö†Ô∏è Pas de donn√©es disponibles")
        return
    
    latest = analyzer.historical_data.iloc[-1]
    
    # Sc√©narios de choc
    scenarios = {
        "Hausse parall√®le +100bps": 0.01,
        "Hausse parall√®le +200bps": 0.02,
        "Baisse parall√®le -50bps": -0.005,
        "Aplatissement (courte +50, longue -50)": None,  # Sp√©cial
        "Pentification (courte -50, longue +50)": None   # Sp√©cial
    }
    
    print(f"Date de r√©f√©rence : {latest['reference_date'].strftime('%Y-%m-%d')}\n")
    print("Taux de base :")
    
    base_rates = {}
    for maturity in [1, 10, 30]:
        col = f'rate_{maturity}y'
        if pd.notna(latest[col]):
            base_rates[maturity] = float(latest[col])
            print(f"  {maturity:2d}Y : {format_rate_pct(base_rates[maturity])}")
    
    print("\n" + "-" * 60)
    print("Sc√©narios de stress :")
    print("-" * 60)
    
    for scenario_name, shock in scenarios.items():
        print(f"\n{scenario_name} :")
        
        for maturity in [1, 10, 30]:
            if maturity in base_rates:
                base = base_rates[maturity]
                
                if shock is not None:
                    # Choc parall√®le
                    stressed = base + shock
                else:
                    # Chocs sp√©ciaux
                    if "Aplatissement" in scenario_name:
                        stressed = base + 0.005 if maturity == 1 else base - 0.005
                    else:  # Pentification
                        stressed = base - 0.005 if maturity == 1 else base + 0.005
                
                change = (stressed - base) * 10000
                print(f"  {maturity:2d}Y : {format_rate_pct(base)} ‚Üí {format_rate_pct(stressed)} ({format_bps(change)})")


def run_all_examples():
    """Ex√©cute tous les exemples"""
    examples = [
        example_1_basic_usage,
        example_2_historical_analysis,
        example_3_multi_maturity_comparison,
        example_4_custom_alerts,
        example_5_export_to_excel,
        example_6_calculate_duration,
        example_7_stress_testing
    ]
    
    print("\n" + "üöÄ" * 40)
    print("EX√âCUTION DE TOUS LES EXEMPLES")
    print("üöÄ" * 40)
    
    for i, example in enumerate(examples, 1):
        try:
            example()
            input(f"\n‚ñ∂Ô∏è  Appuyez sur Entr√©e pour continuer vers l'exemple {i+1}..." 
                  if i < len(examples) else "\n‚úÖ Tous les exemples termin√©s. Appuyez sur Entr√©e...")
        except Exception as e:
            print(f"\n‚ùå Erreur dans l'exemple {i} : {e}")


if __name__ == "__main__":
    import sys
    
    if len(sys.argv) > 1:
        example_num = sys.argv[1]
        
        examples_map = {
            '1': example_1_basic_usage,
            '2': example_2_historical_analysis,
            '3': example_3_multi_maturity_comparison,
            '4': example_4_custom_alerts,
            '5': example_5_export_to_excel,
            '6': example_6_calculate_duration,
            '7': example_7_stress_testing,
            'all': run_all_examples
        }
        
        if example_num in examples_map:
            examples_map[example_num]()
        else:
            print("Usage : python examples.py [1-7|all]")
    else:
        run_all_examples()